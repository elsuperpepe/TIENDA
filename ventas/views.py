import csv

from django.db import transaction
from django.db.models import Sum
from django.http import HttpResponse
from django.shortcuts import render
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from rest_framework import status, viewsets
from rest_framework.response import Response
from rest_framework.views import APIView

from inventario.models import Producto, Tienda
from inventario.serializers import TiendaSerializer

from .models import ItemVenta, Venta
from .serializers import VentaEntradaSerializer
from .services import VentaService, VozParser


class TiendaViewSet(viewsets.ModelViewSet):
    queryset = Tienda.objects.all()
    serializer_class = TiendaSerializer


def index(request):
    return render(request, "index.html")


def pizzeria(request):
    return render(request, "pizzas.html")


class RegistrarVenta(APIView):
    def post(self, request):
        serializer = VentaEntradaSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        datos = serializer.validated_data
        tienda_id = datos["tienda_id"].id
        items_data = [
            {
                "producto_id": item["producto_id"].id,
                "cantidad": item["cantidad"],
                "nota": item.get("nota", ""),
            }
            for item in datos["items"]
        ]
        mesa = request.data.get("mesa", "")
        pago = request.data.get("pago", "efectivo")

        try:
            venta = VentaService.registrar(tienda_id, items_data, mesa=mesa, pago=pago)
            return Response(
                {
                    "mensaje": "Venta realizada con éxito",
                    "venta_id": venta.id,
                    "total": float(venta.total),
                },
                status=status.HTTP_201_CREATED,
            )
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response(
                {"error": "Error inesperado al procesar la venta", "detalle": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class AnularVenta(APIView):
    def post(self, request):
        venta_id = request.data.get("venta_id")
        if not venta_id:
            return Response(
                {"error": "Se requiere venta_id"}, status=status.HTTP_400_BAD_REQUEST
            )
        try:
            venta = VentaService.void(venta_id)
            return Response(
                {
                    "mensaje": f"Venta #{venta.id} anulada correctamente",
                    "total_devuelto": float(venta.total),
                }
            )
        except Venta.DoesNotExist:
            return Response(
                {"error": "Venta no encontrada"}, status=status.HTTP_404_NOT_FOUND
            )
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response(
                {"error": "Error interno", "detalle": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class CerrarDia(APIView):
    def post(self, request):
        ventas_activas = Venta.objects.filter(cancelada=False, cerrada=False)
        total_ventas = ventas_activas.count()

        if total_ventas == 0:
            return Response(
                {"error": "No hay ventas activas para cerrar."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        with transaction.atomic():
            ventas_activas.update(cerrada=True)

        return Response(
            {
                "mensaje": f"Día cerrado correctamente. {total_ventas} venta(s) cerradas.",
                "ventas_cerradas": total_ventas,
            },
            status=status.HTTP_200_OK,
        )


class VentaPorVoz(APIView):
    def post(self, request):
        texto = request.data.get("texto", "").strip()
        tienda_id = request.data.get("tienda_id", 1)

        if not texto:
            return Response(
                {"error": "Envía un texto con tu pedido."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            items = VozParser.parse(texto)
            venta = VentaService.registrar(
                tienda_id, items, mesa="Voz", pago="efectivo"
            )
            return Response(
                {
                    "mensaje": "Venta por voz exitosa",
                    "texto_recibido": texto,
                    "venta_id": venta.id,
                    "total": float(venta.total),
                    "items_detectados": items,
                },
                status=status.HTTP_201_CREATED,
            )
        except ValueError as e:
            return Response({"error": str(e)}, status=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return Response(
                {"error": "Error interno", "detalle": str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )


class ListarVentas(APIView):
    def get(self, request):
        ventas = Venta.objects.filter(cancelada=False, cerrada=False)
        lista = [
            {
                "id": v.id,
                "total": float(v.total),
                "mesa": v.mesa,
                "pago": v.pago,
            }
            for v in ventas
        ]
        total_general = ventas.aggregate(Sum("total"))["total__sum"] or 0
        return Response({"ventas": lista, "total_general": float(total_general)})


class ExportarVentas(APIView):
    def get(self, request):
        wb = Workbook()
        ws = wb.active
        ws.title = "Ventas del día"

        headers = ["#", "Fecha", "Mesa", "Pago", "Items", "Total", "Estado"]
        ws.append(headers)
        for col, cell in enumerate(ws[1], 1):
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="E8250A")
            cell.alignment = Alignment(horizontal="center")

        ventas_qs = Venta.objects.filter(
            cerrada=False, cancelada=False
        ).prefetch_related("items__producto")

        for venta in ventas_qs:
            items_texto = ", ".join(
                f"{i.cantidad}x {i.producto.nombre}"
                + (f" ({i.nota})" if i.nota else "")
                for i in venta.items.all()
            )
            ws.append(
                [
                    venta.id,
                    venta.fecha.strftime("%d/%m/%Y %H:%M"),
                    venta.mesa or "—",
                    venta.pago.capitalize(),
                    items_texto,
                    float(venta.total),
                    "Activa",
                ]
            )

        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

        ws.append([])
        total_row = ws.max_row + 1
        total_general = sum(float(v.total) for v in ventas_qs)
        ws.cell(total_row, 5, "TOTAL DEL DÍA").font = Font(bold=True)
        ws.cell(total_row, 6, total_general).font = Font(bold=True, color="22C55E")

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = (
            'attachment; filename="ventas_don_franco.xlsx"'
        )
        wb.save(response)
        return response


class DetalleVenta(APIView):
    def get(self, request, venta_id):
        try:
            venta = Venta.objects.get(id=venta_id)
        except Venta.DoesNotExist:
            return Response(
                {"error": "Venta no encontrada"}, status=status.HTTP_404_NOT_FOUND
            )

        items_data = []
        for item in venta.items.all():
            items_data.append(
                {
                    "producto_id": item.producto.id,
                    "producto": item.producto.nombre,
                    "nombre": item.producto.nombre,
                    "cantidad": item.cantidad,
                    "precio_unitario": float(item.precio_unitario),
                    "subtotal": float(item.precio_unitario) * item.cantidad,
                    "nota": item.nota,
                }
            )

        return Response(
            {
                "venta_id": venta.id,
                "tienda": venta.tienda.nombre,
                "fecha": venta.fecha.strftime("%Y-%m-%d %H:%M"),
                "total": float(venta.total),
                "cancelada": venta.cancelada,
                "cerrada": venta.cerrada,
                "items": items_data,
            }
        )


class EditarVenta(APIView):
    def post(self, request, venta_id):
        try:
            venta = Venta.objects.prefetch_related("items__producto").get(
                id=venta_id, cancelada=False, cerrada=False
            )
        except Venta.DoesNotExist:
            return Response({"error": "Venta no encontrada o no editable"}, status=404)

        with transaction.atomic():
            # Actualizar mesa y pago si vienen
            venta.mesa = request.data.get("mesa", venta.mesa)
            venta.pago = request.data.get("pago", venta.pago)

            # Si vienen items nuevos, reemplazar todo
            items_nuevos = request.data.get("items")
            if items_nuevos is not None:
                # Restaurar stock de items anteriores
                for item in venta.items.all():
                    if not item.producto.es_servicio:
                        item.producto.stock += item.cantidad
                        item.producto.save()
                venta.items.all().delete()

                total = 0
                for it in items_nuevos:
                    producto = Producto.objects.select_for_update().get(
                        id=it["producto_id"]
                    )
                    cantidad = int(it["cantidad"])
                    nota = it.get("nota", "")
                    if not producto.es_servicio and producto.stock < cantidad:
                        raise ValueError(f"Stock insuficiente para '{producto.nombre}'")
                    ItemVenta.objects.create(
                        venta=venta,
                        producto=producto,
                        cantidad=cantidad,
                        precio_unitario=producto.precio,
                        nota=nota,
                    )
                    if not producto.es_servicio:
                        producto.stock -= cantidad
                        producto.save()
                    total += producto.precio * cantidad

                venta.total = total

            venta.save()

        return Response(
            {
                "mensaje": f"Venta #{venta.id} actualizada",
                "venta_id": venta.id,
                "total": float(venta.total),
            }
        )
